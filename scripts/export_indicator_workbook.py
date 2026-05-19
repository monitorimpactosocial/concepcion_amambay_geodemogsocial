#!/usr/bin/env python3
"""Export the app indicator series to an indexed Excel workbook."""

from __future__ import annotations

import argparse
import json
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Falta openpyxl. Instalarlo o ejecutar en un entorno que lo tenga disponible."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "exports" / "monitor_indicadores_series.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(italic=True, color="666666")


def collect_payload() -> dict[str, Any]:
    script = ROOT / "scripts" / "collect_indicator_series.cjs"
    result = subprocess.run(
        ["node", str(script)],
        cwd=ROOT,
        check=True,
        encoding="utf-8",
        capture_output=True,
    )
    return json.loads(result.stdout)


def safe_sheet_name(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch not in "[]:*?/\\")
    return cleaned[:31]


def append_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200])
        width = max(10, min(42, max(max_len + 2, len(header) + 2)))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if ws.max_row >= 2 and ws.max_column >= 2:
        ref = ws.dimensions
        table_name = f"T_{safe_sheet_name(ws.title).replace('-', '_')}"
        table_name = "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in table_name)
        table = Table(displayName=table_name[:30], ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def build_index_rows(payload: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for sheet_name, spec in payload["sheets"].items():
        rows.append([
            sheet_name,
            spec.get("category", ""),
            spec.get("description", ""),
            len(spec.get("rows", [])),
            len(spec.get("headers", [])),
            spec.get("period", ""),
            spec.get("unit", ""),
            spec.get("source", ""),
        ])
    return rows


def build_sources_rows(payload: dict[str, Any]) -> list[list[Any]]:
    sources: dict[str, set[str]] = {}
    for sheet_name, spec in payload["sheets"].items():
        source = str(spec.get("source", "")).strip()
        if source:
            sources.setdefault(source, set()).add(sheet_name)
        headers = spec.get("headers", [])
        rows = spec.get("rows", [])
        for source_col in ("fuente", "source"):
            if source_col in headers:
                idx = headers.index(source_col)
                for row in rows:
                    if idx < len(row) and row[idx]:
                        sources.setdefault(str(row[idx]).strip(), set()).add(sheet_name)

    return [[source, ", ".join(sorted(sheets))] for source, sheets in sorted(sources.items())]


def write_metadata(ws, payload: dict[str, Any]) -> None:
    ws["A1"] = payload.get("app", "Monitor")
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Generado"
    ws["B2"] = payload.get("generatedAt", "")
    ws["A3"] = "Archivo"
    ws["B3"] = "monitor_indicadores_series.xlsx"
    ws["A5"] = "Notas"
    ws["A5"].font = Font(bold=True)
    for offset, note in enumerate(payload.get("notes", []), start=6):
        ws[f"A{offset}"] = note
        ws[f"A{offset}"].font = NOTE_FONT

    bcp = payload.get("bcpAnexo", {})
    start = 10 + len(payload.get("notes", []))
    ws[f"A{start}"] = "Referencia BCP"
    ws[f"A{start}"].font = Font(bold=True)
    for i, key in enumerate(["title", "fileName", "publishedDate", "sourcePage", "downloadUrl", "note"], start=start + 1):
        ws[f"A{i}"] = key
        ws[f"B{i}"] = bcp.get(key, "")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(payload: dict[str, Any], output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    metadata = wb.create_sheet("METADATOS")
    write_metadata(metadata, payload)

    index = wb.create_sheet("INDICE")
    append_rows(
        index,
        ["hoja", "bloque", "contenido", "filas", "columnas", "periodo", "unidad", "fuente_principal"],
        build_index_rows(payload),
    )
    index.sheet_view.showGridLines = False

    sources = wb.create_sheet("FUENTES")
    append_rows(sources, ["fuente", "hojas_donde_aparece"], build_sources_rows(payload))
    sources.sheet_view.showGridLines = False

    for sheet_name, spec in payload["sheets"].items():
        ws = wb.create_sheet(safe_sheet_name(sheet_name))
        append_rows(ws, spec["headers"], spec["rows"])
        ws.sheet_view.showGridLines = False
        ws["A1"].fill = SUBTLE_FILL
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.properties.creator = "Monitor de Impacto Social"
    wb.properties.title = "Indicadores y series del monitor"
    wb.properties.subject = "Concepcion, Amambay, PARACEL"
    wb.properties.created = datetime.now()
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Ruta del .xlsx a generar")
    args = parser.parse_args()

    payload = collect_payload()
    build_workbook(payload, args.output)
    total_rows = sum(len(spec["rows"]) for spec in payload["sheets"].values())
    print(f"Excel generado: {args.output}")
    print(f"Hojas de datos: {len(payload['sheets'])}")
    print(f"Filas exportadas: {total_rows}")


if __name__ == "__main__":
    main()
